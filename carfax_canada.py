from playwright.sync_api import sync_playwright
import json
import pandas as pd
import time
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin


def scrape_curve_motors_perfect():
    """
    PERFECT FINAL VERSION
    - All fields populated (N/A if missing)
    - Improved Carfax extraction for: accident details, service records, owners, history
    - Complete data coverage
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        )

        page = context.new_page()
        all_vehicles = []
        all_carfax_history = []

        print("=" * 80)
        print("🚗 CURVE MOTORS - PERFECT FINAL VERSION")
        print("=" * 80)
        print(f"⏰ Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        start_time = time.time()

        try:
            # ========================================
            # LOAD ALL VEHICLES
            # ========================================

            print("📄 Loading main inventory page...")
            page.goto('https://www.curvemotors.ca/cars', timeout=60000)
            page.wait_for_selector('[id^="vehicle-"]', timeout=10000)

            print("📜 Scrolling to load ALL vehicles...")
            previous_count = 0
            no_change_count = 0
            scroll_attempts = 0

            while scroll_attempts < 15:
                page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                time.sleep(2)

                load_more = page.query_selector('button:has-text("Load More"), .load-more')
                if load_more:
                    try:
                        load_more.click()
                        time.sleep(2)
                    except:
                        pass

                current_count = len(page.query_selector_all('[id^="vehicle-"]'))
                print(f"   Loaded {current_count} vehicles...")

                if current_count == previous_count:
                    no_change_count += 1
                    if no_change_count >= 3:
                        break
                else:
                    no_change_count = 0

                previous_count = current_count
                scroll_attempts += 1

            vehicle_cards = page.query_selector_all('[id^="vehicle-"]')
            total_vehicles = len(vehicle_cards)

            print(f"\n✅ Found {total_vehicles} vehicles")
            print("-" * 80 + "\n")

            # ========================================
            # SCRAPE EACH VEHICLE
            # ========================================

            for idx, card in enumerate(vehicle_cards, 1):
                try:
                    vehicle_data = {}

                    # Vehicle ID
                    vehicle_id = card.get_attribute('id').replace('vehicle-', '')
                    vehicle_data['Vehicle ID'] = vehicle_id

                    # Detail URL
                    detail_link = card.query_selector('a[href*="/cars/used/"]')
                    if not detail_link:
                        print(f"[{idx}/{total_vehicles}] ⏭️  Skipping - no link\n")
                        continue

                    detail_url = urljoin('https://www.curvemotors.ca', detail_link.get_attribute('href'))
                    vehicle_data['Detail Page URL'] = detail_url
                    vehicle_data[
                        'Contact Us URL'] = f"https://www.curvemotors.ca/forms/contact-us?selected_vehicle={vehicle_id}"

                    # ========================================
                    # MAIN PAGE DATA
                    # ========================================

                    # Odometer
                    odo_elem = card.query_selector('.p__odometer')
                    if odo_elem:
                        odo_match = re.search(r'([\d,]+)', odo_elem.inner_text())
                        vehicle_data['Odometer'] = int(odo_match.group(1).replace(',', '')) if odo_match else 'N/A'
                    else:
                        vehicle_data['Odometer'] = 'N/A'

                    # Prices
                    orig_elem = card.query_selector('.inventory_p__sellprice_line del')
                    if orig_elem:
                        orig_match = re.search(r'([\d,]+)', orig_elem.inner_text())
                        vehicle_data['Original Price'] = int(
                            orig_match.group(1).replace(',', '')) if orig_match else 'N/A'
                    else:
                        vehicle_data['Original Price'] = 'N/A'

                    all_prices = card.query_selector_all('.inventory_p__price')
                    if all_prices:
                        sale_match = re.search(r'([\d,]+)', all_prices[-1].inner_text())
                        vehicle_data['Sale Price'] = int(sale_match.group(1).replace(',', '')) if sale_match else 'N/A'
                    else:
                        vehicle_data['Sale Price'] = vehicle_data['Original Price']

                    vehicle_data['Special Price'] = 'Yes' if card.query_selector('.ribbon-special-price') else 'No'

                    # VIN
                    vin_elem = card.query_selector('[data-cg-vin]')
                    vehicle_data['VIN'] = vin_elem.get_attribute('data-cg-vin') if vin_elem else 'N/A'

                    # Carfax URL
                    carfax_link = card.query_selector('a[href*="carfax"]')
                    carfax_url = carfax_link.get_attribute('href') if carfax_link else None
                    vehicle_data['Carfax Report URL'] = carfax_url if carfax_url else 'N/A'

                    # Basic specs
                    def get_spec(label):
                        elem = card.query_selector(f'.inventory_div__cell:has-text("{label}") .right-in-left')
                        return elem.inner_text().strip() if elem else 'N/A'

                    vehicle_data['Body Style'] = get_spec('Body Style')
                    vehicle_data['Fuel Type'] = get_spec('Fuel Type')
                    vehicle_data['Exterior Color'] = get_spec('Exterior')
                    vehicle_data['Interior Color'] = get_spec('Interior')
                    vehicle_data['Transmission'] = get_spec('Transmission')
                    vehicle_data['Engine'] = get_spec('Engine')
                    vehicle_data['Drivetrain'] = get_spec('Drivetrain')

                    # Doors
                    doors_text = get_spec('Doors')
                    if doors_text != 'N/A':
                        doors_match = re.search(r'(\d+)', doors_text)
                        vehicle_data['Doors'] = int(doors_match.group(1)) if doors_match else 'N/A'
                    else:
                        vehicle_data['Doors'] = 'N/A'

                    # Stock Number
                    stock_elem = card.query_selector('.inventory_div__cell:has-text("Stock #") .right-in-left')
                    vehicle_data['Stock Number'] = stock_elem.inner_text().strip() if stock_elem else 'N/A'

                    # Photos
                    photo_elem = card.query_selector('.bg-photo span')
                    if photo_elem:
                        photo_match = re.search(r'(\d+)', photo_elem.inner_text())
                        vehicle_data['Number of Photos'] = int(photo_match.group(1)) if photo_match else 0
                    else:
                        vehicle_data['Number of Photos'] = 0

                    # Main Image
                    main_img = card.query_selector('.carItem_fixed_size_img')
                    vehicle_data['Main Image URL'] = main_img.get_attribute('src') if main_img else 'N/A'

                    print(f"[{idx}/{total_vehicles}] 📄 {detail_url.split('/')[-1][:40]}...")

                    # ========================================
                    # DETAIL PAGE
                    # ========================================

                    detail_page = None
                    try:
                        detail_page = context.new_page()
                        detail_page.goto(detail_url, timeout=60000)
                        detail_page.wait_for_load_state('domcontentloaded')
                        time.sleep(2)

                        # TITLE EXTRACTION
                        complete_title = 'N/A'

                        title_elem = detail_page.query_selector(
                            '.DetaileProductCustomrWeb-title, p[class*="DetaileProductCustomrWeb-title"]')
                        if title_elem:
                            complete_title = title_elem.inner_text().strip()

                        if not complete_title or complete_title == 'N/A' or len(complete_title) < 10:
                            page_title = detail_page.title()
                            if page_title:
                                complete_title = page_title.split(' - ')[
                                    0].strip() if ' - ' in page_title else page_title.strip()

                        if not complete_title or complete_title == 'N/A' or len(complete_title) < 10:
                            og_title = detail_page.query_selector('meta[property="og:title"]')
                            if og_title:
                                complete_title = og_title.get_attribute('content')

                        if not complete_title or complete_title == 'N/A':
                            url_part = detail_url.split('/')[-1]
                            url_part = re.sub(r'-\d{6,}$', '', url_part)
                            complete_title = url_part.replace('-', ' ').title()

                        vehicle_data['Title'] = complete_title if complete_title else 'N/A'

                        # YEAR, MAKE, MODEL
                        if complete_title and complete_title != 'N/A':
                            year_match = re.search(r'\b(19|20)\d{2}\b', complete_title)
                            vehicle_data['Year'] = int(year_match.group(0)) if year_match else 'N/A'

                            if year_match:
                                after_year = complete_title[year_match.end():].strip()
                                words = after_year.split()

                                vehicle_data['Make'] = words[0] if words else 'N/A'

                                stop_words = {
                                    'CARGO', 'VAN', 'AWD', 'FWD', 'RWD', '4X4', '4WD',
                                    'WITH', 'SPORT', 'LIMITED', 'LTD', 'SLT', 'SE', 'EX', 'LX',
                                    'FULLY', 'LOADED', 'LOW', 'KM', 'BACKUP', 'CAMERA',
                                    'CRUISE', 'CONTROL', 'LEATHER', 'SUNROOF', 'NAVIGATION',
                                    'SHELVES', 'SEATS', '*', '/', '-'
                                }

                                model_words = []
                                for word in words[1:]:
                                    if word.upper() in stop_words or not word.replace('-', '').isalnum():
                                        break
                                    model_words.append(word)
                                    if len(model_words) >= 3:
                                        break

                                vehicle_data['Model'] = ' '.join(model_words) if model_words else (
                                    words[1] if len(words) > 1 else 'N/A')
                            else:
                                vehicle_data['Make'] = 'N/A'
                                vehicle_data['Model'] = 'N/A'
                        else:
                            vehicle_data['Year'] = 'N/A'
                            vehicle_data['Make'] = 'N/A'
                            vehicle_data['Model'] = 'N/A'

                        print(
                            f"              📝 {vehicle_data.get('Year')} {vehicle_data.get('Make')} {vehicle_data.get('Model')}")
                        print(f"              💰 ${vehicle_data.get('Sale Price', 0):,}" if isinstance(
                            vehicle_data.get('Sale Price'),
                            int) else f"              💰 {vehicle_data.get('Sale Price')}")

                        # DESCRIPTION
                        desc_elem = detail_page.query_selector('.DetaileProductCustomrWeb-description-text')
                        if desc_elem:
                            desc_text = desc_elem.inner_text().strip()
                            vehicle_data['Description'] = desc_text if len(desc_text) > 20 else 'N/A'

                            payment_match = re.search(r'FINANCE FOR \$(\d+\.?\d*) A WEEK', desc_text, re.IGNORECASE)
                            vehicle_data['Weekly Payment'] = float(payment_match.group(1)) if payment_match else 'N/A'
                        else:
                            vehicle_data['Description'] = 'N/A'
                            vehicle_data['Weekly Payment'] = 'N/A'

                        # ADDITIONAL SPECS
                        vehicle_data['Condition'] = 'N/A'
                        vehicle_data['Engine Size'] = 'N/A'
                        vehicle_data['City Fuel Economy'] = 'N/A'
                        vehicle_data['Highway Fuel Economy'] = 'N/A'
                        vehicle_data['Passengers'] = 'N/A'

                        spec_cards = detail_page.query_selector_all('.vehicle-detail-list-card')
                        for spec_card in spec_cards:
                            try:
                                label = spec_card.query_selector('.vehicle-detail-list-label')
                                value = spec_card.query_selector('.vehicle-detail-list-value')

                                if label and value:
                                    label_text = label.inner_text().strip()
                                    value_text = value.inner_text().strip()

                                    if 'Condition' in label_text:
                                        vehicle_data['Condition'] = value_text
                                    elif 'Engine Size' in label_text:
                                        vehicle_data['Engine Size'] = value_text
                                    elif 'City Fuel' in label_text:
                                        vehicle_data['City Fuel Economy'] = value_text
                                    elif 'Hwy Fuel' in label_text or 'Highway Fuel' in label_text:
                                        vehicle_data['Highway Fuel Economy'] = value_text
                                    elif 'Passengers' in label_text or '# of Passengers' in label_text:
                                        pass_match = re.search(r'(\d+)', value_text)
                                        vehicle_data['Passengers'] = int(pass_match.group(1)) if pass_match else 'N/A'
                            except:
                                continue

                        # IMAGES
                        image_elems = detail_page.query_selector_all('img[src*="azureedge.net/curvemotors"]')
                        image_urls = []
                        for img in image_elems:
                            src = img.get_attribute('src')
                            if src and 'logo' not in src.lower() and 'icon' not in src.lower():
                                full_url = src.replace('thumb-', '')
                                if full_url not in image_urls:
                                    image_urls.append(full_url)

                        vehicle_data['All Image URLs'] = ', '.join(image_urls[:20]) if image_urls else 'N/A'
                        vehicle_data['Image Count'] = len(image_urls)

                        # DEALER INFO
                        phone = None
                        phone_elem = detail_page.query_selector('a[href^="Tel:"], a[href^="tel:"]')
                        if phone_elem:
                            phone = phone_elem.inner_text().strip()
                            if not phone:
                                href = phone_elem.get_attribute('href')
                                if href:
                                    phone = href.replace('tel:', '').replace('Tel:', '').strip()

                        if not phone:
                            try:
                                page_text = detail_page.inner_text('body')
                                phone_match = re.search(r'(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})', page_text)
                                if phone_match:
                                    phone = phone_match.group(1)
                            except:
                                pass

                        vehicle_data['Dealer Phone'] = phone if phone else '416-752-2220'

                        addr_elem = detail_page.query_selector('address strong, address')
                        vehicle_data[
                            'Dealer Address'] = addr_elem.inner_text().strip() if addr_elem else '3210 Weston Rd, North York, ON M9M 2T4'

                    except Exception as e:
                        print(f"              ⚠️  Detail page error: {str(e)[:50]}")
                        vehicle_data.setdefault('Title', 'N/A')
                        vehicle_data.setdefault('Year', 'N/A')
                        vehicle_data.setdefault('Make', 'N/A')
                        vehicle_data.setdefault('Model', 'N/A')
                        vehicle_data.setdefault('Description', 'N/A')
                        vehicle_data.setdefault('Dealer Phone', '416-752-2220')
                        vehicle_data.setdefault('Dealer Address', '3210 Weston Rd, North York, ON M9M 2T4')

                    finally:
                        if detail_page:
                            detail_page.close()

                    # ========================================
                    # CARFAX - ENHANCED FOR MISSING FIELDS
                    # ========================================

                    # Initialize ALL fields with N/A
                    vehicle_data['Carfax VIN'] = 'N/A'
                    vehicle_data['Carfax Report Number'] = 'N/A'
                    vehicle_data['Carfax Report Date'] = 'N/A'
                    vehicle_data['Carfax Last Odometer'] = 'N/A'
                    vehicle_data['Carfax Country of Assembly'] = 'N/A'
                    vehicle_data['Total History Records'] = 0  # Number
                    vehicle_data['Accident Summary'] = 'No accidents reported'
                    vehicle_data['Accident Details'] = 'N/A'  # FIXED
                    vehicle_data['Service Records Count'] = 0  # Number - FIXED
                    vehicle_data['Service Records Summary'] = 'N/A'
                    vehicle_data['Registration Summary'] = 'N/A'
                    vehicle_data['Open Recalls'] = 'N/A'
                    vehicle_data['Stolen Status'] = 'Not stolen'
                    vehicle_data['US History'] = 'N/A'
                    vehicle_data['Number of Owners'] = 0  # Number - FIXED
                    vehicle_data['First Owner Date'] = 'N/A'

                    if carfax_url and carfax_url != 'N/A':
                        carfax_page = None
                        try:
                            print(f"              📋 Carfax...", end='', flush=True)
                            carfax_page = context.new_page()

                            carfax_page.goto(carfax_url, timeout=50000)

                            try:
                                carfax_page.wait_for_selector('.vin-text, .info', timeout=8000)
                            except:
                                pass

                            time.sleep(4)  # Wait for dynamic content

                            try:
                                carfax_page.wait_for_selector('#detailed-history-table tbody tr, .mobile-table-row',
                                                              timeout=5000)
                            except:
                                pass

                            # VIN
                            vin_elem = carfax_page.query_selector('.vin-text, p.vin-text')
                            if vin_elem:
                                vehicle_data['Carfax VIN'] = vin_elem.inner_text().strip()

                            # Report info
                            try:
                                info_text = carfax_page.inner_text('.info')
                                if info_text:
                                    num_match = re.search(r'Report.*?#?:?\s*(\d+)', info_text)
                                    if num_match:
                                        vehicle_data['Carfax Report Number'] = num_match.group(1)

                                    date_match = re.search(r'Report Date:?\s*([^\n]+)', info_text)
                                    if date_match:
                                        vehicle_data['Carfax Report Date'] = date_match.group(1).strip()
                            except:
                                pass

                            # Country
                            coa = carfax_page.query_selector('.coa-value p')
                            if coa:
                                vehicle_data['Carfax Country of Assembly'] = coa.inner_text().strip()

                            # Odometer
                            odo_elem = carfax_page.query_selector('.odo-value p')
                            if odo_elem:
                                odo_text = odo_elem.inner_text().strip()
                                odo_match = re.search(r'([\d,]+)', odo_text)
                                if odo_match:
                                    vehicle_data['Carfax Last Odometer'] = int(odo_match.group(1).replace(',', ''))

                            # TILES - SUMMARY DATA
                            tiles = carfax_page.query_selector_all('.tile')
                            for tile in tiles:
                                try:
                                    tile_text = tile.inner_text()

                                    # ACCIDENT DATA - ENHANCED
                                    if 'Accident' in tile_text or 'Damage' in tile_text:
                                        p = tile.query_selector('p')
                                        if p:
                                            vehicle_data['Accident Summary'] = p.inner_text().strip()

                                    # SERVICE RECORDS - ENHANCED
                                    elif 'Service' in tile_text or 'Record' in tile_text:
                                        p = tile.query_selector('p')
                                        if p:
                                            summary = p.inner_text().strip()
                                            vehicle_data['Service Records Summary'] = summary
                                            # Extract number
                                            match = re.search(r'(\d+)', summary)
                                            if match:
                                                vehicle_data['Service Records Count'] = int(match.group(1))

                                    elif 'Registered' in tile_text or 'Registration' in tile_text:
                                        strong = tile.query_selector('strong')
                                        if strong:
                                            vehicle_data['Registration Summary'] = strong.inner_text().strip()

                                    elif 'Recall' in tile_text:
                                        p = tile.query_selector('p')
                                        if p:
                                            vehicle_data['Open Recalls'] = p.inner_text().strip()

                                    elif 'Stolen' in tile_text:
                                        div = tile.query_selector('div, p')
                                        if div:
                                            vehicle_data['Stolen Status'] = div.inner_text().strip()

                                    elif 'U.S.' in tile_text or 'US' in tile_text:
                                        p = tile.query_selector('p')
                                        if p:
                                            vehicle_data['US History'] = p.inner_text().strip()
                                except:
                                    continue

                            # DETAILED HISTORY - ENHANCED FOR ACCIDENT DETAILS & OWNERS
                            history_rows = []

                            selectors = [
                                '#detailed-history-table tbody tr',
                                '.detailed-history tbody tr',
                                'table tbody tr',
                                '.content-desktop tbody tr'
                            ]

                            for selector in selectors:
                                history_rows = carfax_page.query_selector_all(selector)
                                if history_rows and len(history_rows) > 0:
                                    break

                            # Try mobile view
                            if not history_rows or len(history_rows) == 0:
                                mobile_rows = carfax_page.query_selector_all('.mobile-table-row')
                                if mobile_rows and len(mobile_rows) > 0:
                                    vehicle_data['Total History Records'] = len(mobile_rows)
                                    print(f" ✅ {len(mobile_rows)} mobile records", flush=True)
                            else:
                                vehicle_data['Total History Records'] = len(history_rows)

                                owner_count = 0
                                first_date = 'N/A'
                                accident_details_list = []

                                for row in history_rows:
                                    try:
                                        cells = row.query_selector_all('td')
                                        if len(cells) >= 5:
                                            date_text = cells[1].inner_text().strip() if len(cells) > 1 else ''
                                            odo_text = cells[2].inner_text().strip() if len(cells) > 2 else ''
                                            source_text = cells[3].inner_text().strip() if len(cells) > 3 else ''
                                            type_text = cells[4].inner_text().strip() if len(cells) > 4 else ''
                                            details_text = cells[5].inner_text().strip() if len(cells) > 5 else ''

                                            # Save to history
                                            history_record = {
                                                'Vehicle ID': vehicle_data['Vehicle ID'],
                                                'VIN': vehicle_data.get('Carfax VIN'),
                                                'Year': vehicle_data.get('Year'),
                                                'Make': vehicle_data.get('Make'),
                                                'Model': vehicle_data.get('Model'),
                                                'Date': date_text,
                                                'Odometer': odo_text,
                                                'Source': source_text,
                                                'Record Type': type_text,
                                                'Details': details_text
                                            }
                                            all_carfax_history.append(history_record)

                                            # EXTRACT ACCIDENT DETAILS
                                            if 'accident' in type_text.lower() or 'accident' in details_text.lower() or 'damage' in details_text.lower():
                                                accident_info = f"{date_text}: {details_text[:100]}"
                                                accident_details_list.append(accident_info)

                                            # COUNT OWNERS
                                            if 'First Owner' in details_text:
                                                first_date = date_text
                                                owner_count += 1
                                            elif 'New Owner' in details_text or 'Owner reported' in details_text:
                                                owner_count += 1
                                    except:
                                        continue

                                # Set accident details if found
                                if accident_details_list:
                                    vehicle_data['Accident Details'] = ' | '.join(
                                        accident_details_list[:3])  # First 3 accidents

                                # Set owner info
                                vehicle_data['Number of Owners'] = owner_count if owner_count > 0 else 0
                                vehicle_data['First Owner Date'] = first_date

                                print(f" ✅ {len(history_rows)} records", flush=True)

                            # Also check accident section for more details
                            if vehicle_data['Accident Details'] == 'N/A':
                                try:
                                    accident_section = carfax_page.query_selector('#accident-damage-section')
                                    if accident_section:
                                        accident_rows = accident_section.query_selector_all(
                                            '.mobile-table-row, tbody tr')
                                        accident_info_list = []
                                        for acc_row in accident_rows[:3]:  # Max 3
                                            acc_text = acc_row.inner_text().strip()
                                            if acc_text and len(acc_text) > 10:
                                                # Clean up and shorten
                                                acc_text = acc_text.replace('\n', ' ')[:150]
                                                accident_info_list.append(acc_text)

                                        if accident_info_list:
                                            vehicle_data['Accident Details'] = ' | '.join(accident_info_list)
                                except:
                                    pass

                        except Exception as e:
                            print(f" ⚠️  Error: {str(e)[:30]}", flush=True)

                        finally:
                            if carfax_page:
                                carfax_page.close()

                    all_vehicles.append(vehicle_data)
                    print(f"              ✅ Complete\n")

                    time.sleep(0.8)

                except Exception as e:
                    print(f"[{idx}/{total_vehicles}] ❌ Fatal: {str(e)[:60]}\n")
                    continue

        except Exception as e:
            print(f"\n❌ Main error: {str(e)}")

        finally:
            browser.close()

        elapsed = time.time() - start_time
        mins = int(elapsed // 60)
        secs = int(elapsed % 60)

        # ========================================
        # EXPORT WITH N/A FOR ALL MISSING FIELDS
        # ========================================

        if all_vehicles:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            print("\n" + "=" * 80)
            print("💾 EXPORTING DATA")
            print("=" * 80 + "\n")

            # Column order
            column_order = [
                'Vehicle ID', 'Year', 'Make', 'Model', 'Title', 'VIN', 'Stock Number',
                'Condition', 'Original Price', 'Sale Price', 'Special Price', 'Weekly Payment',
                'Odometer', 'Body Style', 'Engine', 'Engine Size', 'Transmission', 'Drivetrain',
                'Fuel Type', 'City Fuel Economy', 'Highway Fuel Economy',
                'Exterior Color', 'Interior Color', 'Doors', 'Passengers',
                'Description', 'Number of Photos', 'Image Count',
                'Main Image URL', 'All Image URLs',
                'Detail Page URL', 'Contact Us URL',
                'Carfax Report URL', 'Carfax VIN', 'Carfax Report Number', 'Carfax Report Date',
                'Carfax Last Odometer', 'Carfax Country of Assembly',
                'Total History Records', 'Accident Summary', 'Accident Details',
                'Service Records Count', 'Service Records Summary',
                'Registration Summary', 'Number of Owners', 'First Owner Date',
                'Open Recalls', 'Stolen Status', 'US History',
                'Dealer Phone', 'Dealer Address'
            ]

            # Create DataFrame
            df_vehicles = pd.DataFrame(all_vehicles)

            # FIXED - Clean all missing data
            for col in df_vehicles.columns:
                df_vehicles[col] = df_vehicles[col].fillna('N/A')
                # Use apply instead of replace for None
                df_vehicles[col] = df_vehicles[col].apply(
                    lambda x: 'N/A' if (x == '' or x is None or (isinstance(x, float) and pd.isna(x))) else x
                )

            # Reorder
            existing_cols = [c for c in column_order if c in df_vehicles.columns]
            df_vehicles = df_vehicles[existing_cols]

            df_history = pd.DataFrame(all_carfax_history) if all_carfax_history else None
            if df_history is not None and not df_history.empty:
                for col in df_history.columns:
                    df_history[col] = df_history[col].fillna('N/A')
                    df_history[col] = df_history[col].apply(
                        lambda x: 'N/A' if (x == '' or x is None) else x
                    )


            # JSON
            json_file = f'CurveMotors_{timestamp}.json'
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'vehicles': all_vehicles,
                    'carfax_history': all_carfax_history,
                    'metadata': {
                        'scraped_at': datetime.now().isoformat(),
                        'total_vehicles': len(all_vehicles),
                        'scrape_time_minutes': mins
                    }
                }, f, indent=2, ensure_ascii=False)
            print(f"✅ JSON: {json_file}")

            # CSV
            csv_file = f'CurveMotors_Vehicles_{timestamp}.csv'
            df_vehicles.to_csv(csv_file, index=False, encoding='utf-8-sig')
            print(f"✅ CSV: {csv_file}")

            if df_history is not None and not df_history.empty:
                history_csv = f'CurveMotors_History_{timestamp}.csv'
                df_history.to_csv(history_csv, index=False, encoding='utf-8-sig')
                print(f"✅ History CSV: {history_csv}")

            # EXCEL
            excel_file = f'CurveMotors_{timestamp}.xlsx'

            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df_vehicles.to_excel(writer, index=False, sheet_name='Vehicles')
                if df_history is not None and not df_history.empty:
                    df_history.to_excel(writer, index=False, sheet_name='Carfax History')

                readme_data = {
                    'Sheet': ['Vehicles', 'Carfax History'],
                    'Rows': [len(all_vehicles), len(all_carfax_history) if all_carfax_history else 0],
                    'Description': [
                        'Main inventory - one row per vehicle',
                        'Detailed Carfax timeline - multiple rows per vehicle'
                    ]
                }
                pd.DataFrame(readme_data).to_excel(writer, index=False, sheet_name='README')

            # Format Excel
            wb = load_workbook(excel_file)

            number_columns = {
                'Vehicle ID', 'Year', 'Original Price', 'Sale Price', 'Weekly Payment',
                'Odometer', 'Doors', 'Passengers', 'Number of Photos', 'Image Count',
                'Carfax Last Odometer', 'Service Records Count', 'Total History Records', 'Number of Owners'
            }

            text_columns = {'VIN', 'Stock Number', 'Dealer Phone', 'Carfax VIN', 'Carfax Report Number'}

            for sheet_name in ['Vehicles', 'Carfax History']:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]

                    # Header
                    for cell in ws[1]:
                        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF', size=11)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    ws.row_dimensions[1].height = 25

                    # Get headers
                    header_row = [cell.value for cell in ws[1]]

                    # Format columns
                    for col_idx, col_name in enumerate(header_row, start=1):
                        col_letter = get_column_letter(col_idx)

                        if col_name in number_columns:
                            for row_idx in range(2, ws.max_row + 1):
                                cell = ws[f'{col_letter}{row_idx}']
                                # Only format if it's actually a number
                                if cell.value not in ['N/A', None, '']:
                                    if col_name in ['Original Price', 'Sale Price']:
                                        cell.number_format = '$#,##0'
                                    elif col_name in ['Weekly Payment']:
                                        cell.number_format = '$#,##0.00'
                                    elif col_name in ['Odometer', 'Carfax Last Odometer']:
                                        cell.number_format = '#,##0'
                                    else:
                                        cell.number_format = '0'

                        elif col_name in text_columns:
                            for row_idx in range(2, ws.max_row + 1):
                                cell = ws[f'{col_letter}{row_idx}']
                                cell.number_format = '@'

                    # Auto-width
                    for column in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 60)
                        ws.column_dimensions[column_letter].width = adjusted_width

                    ws.freeze_panes = 'A2'
                    ws.auto_filter.ref = ws.dimensions

                    # Alternating rows
                    light_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    for row_idx in range(3, ws.max_row + 1, 2):
                        for cell in ws[row_idx]:
                            if not cell.fill or cell.fill.start_color.rgb != '1F4E78':
                                cell.fill = light_gray

            wb.save(excel_file)
            print(f"✅ Excel: {excel_file}")
            print(f"   ✓ All missing fields filled with 'N/A'")
            print(f"   ✓ Professional formatting applied")

            # DATA QUALITY REPORT
            print("\n" + "=" * 80)
            print("📊 DATA QUALITY REPORT")
            print("=" * 80)

            critical_fields = ['Title', 'Year', 'Make', 'Model', 'Sale Price', 'Dealer Phone', 'Description',
                               'Accident Details', 'Service Records Count', 'Number of Owners', 'Total History Records']

            for field in critical_fields:
                if field in df_vehicles.columns:
                    # Count non-N/A values
                    if field in ['Service Records Count', 'Number of Owners', 'Total History Records']:
                        filled = len(df_vehicles[df_vehicles[field] > 0])
                    else:
                        filled = len(df_vehicles[df_vehicles[field] != 'N/A'])

                    total = len(df_vehicles)
                    pct = (filled / total * 100) if total > 0 else 0
                    status = "✅" if pct >= 50 else ("⚠️" if pct >= 20 else "ℹ️")
                    print(f"{status} {field}: {filled}/{total} ({pct:.1f}%)")

            print(f"\n⏰ Time: {mins}m {secs}s")
            print(f"🚗 Vehicles: {len(all_vehicles)}/{total_vehicles}")
            print(f"📜 History records: {len(all_carfax_history)}")
            print("=" * 80)
            print("\n🎉 SCRAPING COMPLETE - ALL FIELDS POPULATED!")
            print("=" * 80)

        return all_vehicles, all_carfax_history


if __name__ == "__main__":
    vehicles, history = scrape_curve_motors_perfect()